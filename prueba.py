# import json
#
# from Settings import config_paths as cp
#
# paths = cp.paths()
#
# class Aplicacion():
#     def __init__(self):
#         file_paths = paths.curriculos_path
#         with open(file_paths) as json_file:
#             data = json.load(json_file)
#
#         print(data)
#         data[0]['prueba2']='hello'
#         print(data)
#         with open(file_paths,'w+') as file:
#             json.dump(data, file, indent=4)
#
# def main():
#     mi_app = Aplicacion()
#     return 0
#
# if __name__ == '__main__':
#     main()
from openpyxl import load_workbook

file = load_workbook('Data_base/4_APENDICE 4 CGA ST.xlsx')
sheets = file.sheetnames
sheet = file[sheets[0]]

for index_file in range(10, 8000):
    print(sheet.cell(row=index_file, column=3).value)
    if sheet.cell(row=index_file, column=3).value == None:
        print('bye')
        break

print(sheet.cell(row=8, column=1).value)

