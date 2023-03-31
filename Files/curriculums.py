from openpyxl import load_workbook

def load_pe(name_file):
    print (name_file)
    file = load_workbook(name_file)
    sheets = file.sheetnames
    sheet = file[sheets[0]]

    for index_file in range(10, 8000):
        if sheet.cell(row=index_file, column=3).value == None:
            break
        else:
            print(sheet.cell(row=index_file, column=5).value[0])

    print(sheet.cell(row=8, column=1).value)

#load_pe('../Data_base/4_APENDICE 4 CGA ST.xlsx')